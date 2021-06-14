import openpyxl

xlsfile = openpyxl.load_workbook('movies.xlsx')
print(xlsfile.sheetnames)
sheet = xlsfile['Sheet1']
print(sheet['B4'].value)
print(sheet['D4'].value)
print(sheet.max_row)
print(sheet.max_column)
#Not existent rows will return None Values
print(sheet['A12'].value)
print(sheet['E1'].value)

#UPDATE XLSX FILES
from openpyxl.comments import Comment
xlsfile = openpyxl.load_workbook('movies.xlsx')
sheet = xlsfile['Sheet1']
print(sheet['D4'].value)
sheet['D4'].comment = Comment('Changed text automatically', 'User')
sheet['B12'] = '=SUM(B2:B11)'
xlsfile.save('movies_comment.xlsx')

#CREATE NEW EXCEL SPREADSHEET
xls_file = openpyxl.Workbook()
xls_file.sheetnames[0]
sheet = xls_file['Sheet']

data = [(225.7, 'Gone with the wind', 'Victor Fleming'),
        (194.4, 'Star Wars', 'George Lucas'),
        (161.0, 'ET: The Extraterrestrial', 'Steven Spielberg')]
for row, (addmission, name, director) in enumerate(data, 1):
    sheet['A{}'.format(row)].value = addmission
    sheet['B{}'.format(row)].value = name
sheet = xls_file.create_sheet('Directors')
print(sheet)

for row, (addmission, name, director) in enumerate(data, 1):
    sheet['A{}'.format(row)].value = director
    sheet['B{}'.format(row)].value = name


sheet = xls_file.create_sheet('Complete')
for row, (addmission, name, director) in enumerate(data, 1):
    sheet['A{}'.format(row)].value = addmission
    sheet['B{}'.format(row)].value = name
    sheet['C{}'.format(row)].value = director
xls_file.save('Movie_Sheets.xlsx')